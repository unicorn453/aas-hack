import argparse
import json
from typing import Any, Dict, List, Set, Tuple

import openpyxl as xl


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    normalized = normalized.replace("_", " ").replace("-", " ")
    return " ".join(normalized.split())


def _header_match_score(headers: List[str]) -> int:
    semantic_tokens = (
        "semantic",
        "semantics",
    )
    value_tokens = (
        "actual value",
        "initial value",
        "value",
    )

    score = 0
    for header in headers:
        if any(token in header for token in semantic_tokens):
            score += 2
        if any(token == header or token in header for token in value_tokens):
            score += 1
    return score


def find_header_row_index(rows: List[Tuple[Any, ...]], max_scan_rows: int = 30) -> int:
    if not rows:
        raise ValueError("Excel file is empty.")

    best_index = 0
    best_score = -1
    scan_limit = min(len(rows), max_scan_rows)

    for idx in range(scan_limit):
        headers = [normalize_header(cell) for cell in rows[idx]]
        score = _header_match_score(headers)
        if score > best_score:
            best_score = score
            best_index = idx

    return best_index


def find_column_indexes(header_cells: List[Any]) -> Dict[str, Any]:
    normalized = [normalize_header(cell) for cell in header_cells]

    semantic_exact = {
        "semantic id",
        "semantic id iri",
        "semantic id irdi",
        "semantic iri",
        "semantic irdi",
        "semantics iri",
        "semantics irdi",
    }
    value_exact = {
        "actual value",
        "initial value",
        "value",
    }

    semantic_indices: List[int] = []
    value_idx = -1

    for idx, header in enumerate(normalized):
        if not header:
            continue

        if header in semantic_exact or (
            "semantic" in header and ("id" in header or "iri" in header or "irdi" in header)
        ):
            semantic_indices.append(idx)

        if value_idx == -1 and (header in value_exact or "initial value" in header or "actual value" in header):
            value_idx = idx

    # Fallback for IDTA-like templates: V/W are semantics (IRI/IRDI), X is initial value.
    if not semantic_indices:
        for fallback_idx in (21, 22):
            if fallback_idx < len(normalized):
                semantic_indices.append(fallback_idx)

    if value_idx == -1 and 23 < len(normalized):
        value_idx = 23

    semantic_indices = sorted(set(semantic_indices))

    if not semantic_indices or value_idx == -1:
        raise ValueError(
            "Could not find required header columns. "
            "Expected semantic columns (e.g. semantic-id-IRI/IRDI) and a value column "
            "(e.g. actual value or initial value)."
        )

    return {"semantic": semantic_indices, "value": value_idx}


def parse_excel_column_ref(column_ref: str) -> int:
    col = column_ref.strip().upper()
    if not col:
        raise ValueError("Empty column reference is not allowed.")

    if col.isdigit():
        value = int(col)
        if value < 1:
            raise ValueError("Numeric column references are 1-based and must be >= 1.")
        return value - 1

    if not col.isalpha():
        raise ValueError(f"Invalid column reference '{column_ref}'.")

    idx = 0
    for char in col:
        idx = idx * 26 + (ord(char) - ord("A") + 1)
    return idx - 1


def parse_semantic_columns_arg(value: str) -> List[int]:
    refs = [part.strip() for part in value.split(",") if part.strip()]
    if not refs:
        raise ValueError("--semantic-columns requires at least one column reference.")
    return [parse_excel_column_ref(ref) for ref in refs]


def load_semantic_map(
    xlsx_path: str,
    sheet_name: str = "",
    semantic_columns: List[int] = None,
    value_column: int = -1,
) -> Dict[str, Any]:
    workbook = xl.load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty.")

        if semantic_columns is None:
            semantic_columns = []

        if not semantic_columns or value_column < 0:
            header_row_idx = find_header_row_index(rows)
            auto_col_idx = find_column_indexes(list(rows[header_row_idx]))

            if not semantic_columns:
                semantic_columns = auto_col_idx["semantic"]
            if value_column < 0:
                value_column = auto_col_idx["value"]
            data_rows = rows[header_row_idx + 1 :]
        else:
            data_rows = rows

        semantic_to_value: Dict[str, Any] = {}
        for row in data_rows:
            if row is None:
                continue

            semantic_raw = None
            for semantic_idx in semantic_columns:
                if semantic_idx < len(row) and row[semantic_idx] is not None:
                    candidate = str(row[semantic_idx]).strip()
                    if candidate:
                        semantic_raw = candidate
                        break

            value_raw = row[value_column] if value_column < len(row) else None

            if semantic_raw is None:
                continue

            semantic_id = str(semantic_raw).strip()
            if not semantic_id:
                continue

            semantic_to_value[semantic_id] = value_raw

        return semantic_to_value
    finally:
        workbook.close()


def parse_maybe_json(value: Any) -> Any:
    if not isinstance(value, str):
        return value

    stripped = value.strip()
    if not stripped:
        return ""

    if stripped[0] in "[{":
        try:
            return json.loads(stripped)
        except json.JSONDecodeError:
            return value
    return value


def get_model_type_name(node: Dict[str, Any]) -> str:
    model_type = node.get("modelType")
    if isinstance(model_type, str):
        return model_type
    if isinstance(model_type, dict):
        return str(model_type.get("name", ""))
    return ""


def set_value_by_model_type(node: Dict[str, Any], new_value: Any) -> None:
    model_type = get_model_type_name(node)

    if model_type == "MultiLanguageProperty":
        if isinstance(node.get("value"), dict) and "langString" in node["value"]:
            node["value"]["langString"] = [{"language": "en", "text": str(new_value)}]
            return

        if isinstance(node.get("value"), list):
            node["value"] = [{"language": "en", "text": str(new_value)}]
            return

        node["value"] = [{"language": "en", "text": str(new_value)}]
        return

    node["value"] = parse_maybe_json(new_value)


def semantic_ids_of_node(node: Dict[str, Any]) -> List[str]:
    semantic_id = node.get("semanticId")
    if not isinstance(semantic_id, dict):
        return []

    keys = semantic_id.get("keys")
    if not isinstance(keys, list):
        return []

    result: List[str] = []
    for key in keys:
        if isinstance(key, dict) and key.get("value") is not None:
            result.append(str(key["value"]).strip())
    return result


def update_values_by_semantic_id(
    node: Any,
    semantic_map: Dict[str, Any],
    used_semantic_ids: Set[str],
) -> int:
    updates = 0

    if isinstance(node, dict):
        node_semantic_ids = semantic_ids_of_node(node)
        for semantic_id in node_semantic_ids:
            if semantic_id in semantic_map and "value" in node:
                set_value_by_model_type(node, semantic_map[semantic_id])
                used_semantic_ids.add(semantic_id)
                updates += 1
                break

        for value in node.values():
            updates += update_values_by_semantic_id(value, semantic_map, used_semantic_ids)

    elif isinstance(node, list):
        for item in node:
            updates += update_values_by_semantic_id(item, semantic_map, used_semantic_ids)

    return updates


def run(
    excel_path: str,
    input_json_path: str,
    output_json_path: str,
    sheet_name: str,
    semantic_columns: List[int],
    value_column: int,
) -> Tuple[int, int]:
    semantic_map = load_semantic_map(excel_path, sheet_name, semantic_columns, value_column)

    with open(input_json_path, "r", encoding="utf-8") as f:
        aas_json = json.load(f)

    used_semantic_ids: Set[str] = set()
    updated_count = update_values_by_semantic_id(aas_json, semantic_map, used_semantic_ids)

    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(aas_json, f, ensure_ascii=False, indent=2)

    missing_count = len(set(semantic_map.keys()) - used_semantic_ids)
    return updated_count, missing_count


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Map Excel Semantic ID + actual value to an existing AAS JSON by semanticId."
    )
    parser.add_argument("excel", help="Path to the source Excel file")
    parser.add_argument("input_json", help="Path to input AAS JSON")
    parser.add_argument("output_json", help="Path to output AAS JSON")
    parser.add_argument(
        "--sheet",
        default="",
        help="Optional sheet name. If omitted, the first sheet is used.",
    )
    parser.add_argument(
        "--semantic-columns",
        default="",
        help="Optional semantic ID column(s), e.g. 'V' or 'V,W'. If omitted, auto-detected.",
    )
    parser.add_argument(
        "--value-column",
        default="",
        help="Optional value column, e.g. 'X'. If omitted, auto-detected.",
    )

    args = parser.parse_args()
    semantic_columns = parse_semantic_columns_arg(args.semantic_columns) if args.semantic_columns else []
    value_column = parse_excel_column_ref(args.value_column) if args.value_column else -1

    updated_count, missing_count = run(
        args.excel,
        args.input_json,
        args.output_json,
        args.sheet,
        semantic_columns,
        value_column,
    )

    print(f"Updated elements: {updated_count}")
    print(f"Semantic IDs not found in JSON: {missing_count}")


if __name__ == "__main__":
    main()
