"""Extract configured company spreadsheets into inspectable JSON files."""

from __future__ import annotations

import argparse
import datetime
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter

from excel_to_aasx.company_config import DEFAULT_COMPANY_CONFIG, load_company_config
from excel_to_aasx.logging import generated


def text_or_empty(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slug(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").lower()
    return normalized or "asset"


def compact_row(row: list[str]) -> list[str]:
    end = len(row)
    while end > 0 and row[end - 1] == "":
        end -= 1
    return row[:end]


def is_header(row: list[str]) -> bool:
    return bool(row) and row[0] == "Element Name (idShort)"


def section_level(title: str) -> int:
    match = re.match(r"^(\d+(?:\.\d+)*)\b", title)
    if not match:
        return 0
    return match.group(1).count(".") + 1


def parse_sheet(sheet_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    parsed_rows: list[dict[str, Any]] = []
    raw_rows: list[dict[str, Any]] = []
    current_header: list[str] | None = None
    sections: list[dict[str, Any]] = []

    for raw_row in rows:
        row_number = raw_row["row"]
        raw = raw_row["values"]
        row = compact_row([text_or_empty(value) for value in raw])
        if not row:
            continue

        raw_rows.append({"row": row_number, "values": row})

        if is_header(row):
            current_header = row
            continue

        maybe_section = row[0]
        level = section_level(maybe_section)
        if level:
            sections = [item for item in sections if item["level"] < level]
            sections.append({"level": level, "title": maybe_section, "row": row_number})
            continue

        if current_header is None:
            continue

        item = row_to_object(row_number, row, current_header)
        if not item:
            continue
        item["sectionPath"] = [section["title"] for section in sections]
        parsed_rows.append(item)

    return {
        "sheet": sheet_name,
        "rowCount": len(rows),
        "parsedRowCount": len(parsed_rows),
        "parsedRows": parsed_rows,
        "rawRows": raw_rows,
    }


def column_name(index: int) -> str:
    return get_column_letter(index + 1)


def cell_row_number(cell_ref: str) -> int:
    digits = "".join(ch for ch in cell_ref if ch.isdigit())
    return int(digits) if digits else 0


def json_scalar(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return str(value).strip()


def cell_object(cell: Cell | MergedCell) -> dict[str, Any]:
    ref = cell.coordinate
    value = json_scalar(cell.value)

    result = {
        "ref": ref,
        "row": cell.row,
        "column": get_column_letter(cell.column),
        "columnIndex": cell.column - 1,
        "value": display_value(cell.value),
    }
    if cell.data_type:
        result["type"] = cell.data_type
    if cell.style_id:
        result["styleIndex"] = cell.style_id
    if value is not None:
        result["rawValue"] = value
    if cell.data_type == "f":
        result["formula"] = str(cell.value or "").removeprefix("=")
    if cell.number_format:
        result["numberFormat"] = cell.number_format
    if cell.hyperlink and cell.hyperlink.target:
        result["hyperlink"] = cell.hyperlink.target
    if cell.comment:
        result["comment"] = {
            "author": cell.comment.author,
            "text": cell.comment.text,
        }
    return result


def row_attributes(row_dimension: Any, row_number: int) -> dict[str, Any]:
    attributes: dict[str, Any] = {"r": str(row_number)}
    if row_dimension.hidden:
        attributes["hidden"] = True
    if row_dimension.height is not None:
        attributes["height"] = row_dimension.height
    if row_dimension.outlineLevel:
        attributes["outlineLevel"] = row_dimension.outlineLevel
    return attributes


def read_complete_sheet(sheet: Any) -> dict[str, Any]:
    dimension = sheet.calculate_dimension()
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0

    complete_rows: list[dict[str, Any]] = []
    all_cells: list[dict[str, Any]] = []
    for row_number in range(1, max_row + 1):
        values = [""] * max_col
        row_cells = []
        for col_number in range(1, max_col + 1):
            cell = sheet.cell(row=row_number, column=col_number)
            if (
                cell.value is None
                and not getattr(cell, "has_style", False)
                and not getattr(cell, "hyperlink", None)
                and not getattr(cell, "comment", None)
            ):
                continue
            item = cell_object(cell)
            values[item["columnIndex"]] = item["value"]
            row_cells.append(item)
            all_cells.append(item)
        complete_rows.append(
            {
                "row": row_number,
                "attributes": row_attributes(sheet.row_dimensions[row_number], row_number),
                "cells": row_cells,
                "values": values,
            }
        )

    return {
        "dimension": dimension,
        "rowCount": max_row,
        "columnCount": max_col,
        "columns": [column_name(index) for index in range(max_col)],
        "columnMetadata": {
            get_column_letter(index): {
                key: value
                for key, value in {
                    "width": sheet.column_dimensions[get_column_letter(index)].width,
                    "hidden": sheet.column_dimensions[get_column_letter(index)].hidden,
                }.items()
                if value not in {None, False}
            }
            for index in range(1, max_col + 1)
        },
        "mergedRanges": [str(item) for item in sheet.merged_cells.ranges],
        "completeRows": complete_rows,
        "cells": all_cells,
    }


def row_to_object(
    row_number: int,
    row: list[str],
    header: list[str],
) -> dict[str, Any] | None:
    if not row or row[0] == "Element Name (idShort)":
        return None

    values = {
        header[index]: value
        for index, value in enumerate(row)
        if index < len(header) and header[index] and value
    }
    if not values:
        return None

    return {
        "row": row_number,
        "idShort": values.get("Element Name (idShort)", ""),
        "obligation": values.get("Obligation", ""),
        "fieldType": values.get("Data Style / Field Type", ""),
        "description": values.get("Description", ""),
        "semanticId": values.get("Universal Semantic ID (Global Key)", ""),
        "exampleValue": values.get("Example Value", ""),
        "actualValue": values.get("Actual Value", ""),
        "allColumns": values,
    }


def extract_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False, rich_text=False)
    sheets = {}
    for sheet in workbook.worksheets:
        complete_sheet = read_complete_sheet(sheet)
        rows = [
            {"row": item["row"], "values": item["values"]}
            for item in complete_sheet["completeRows"]
        ]
        sheets[sheet.title] = {
            **parse_sheet(sheet.title, rows),
            "completeExtraction": complete_sheet,
        }

    return {
        "source": str(path),
        "workbook": path.name,
        "extractionLevel": "complete-xlsx-step-1",
        "extractor": "openpyxl",
        "sheets": sheets,
    }


def write_workbook_outputs(workbook: dict[str, Any], output_dir: Path) -> None:
    workbook_slug = slug(workbook["workbook"].removesuffix(".xlsx"))
    workbook_dir = output_dir / workbook_slug
    workbook_dir.mkdir(parents=True, exist_ok=True)

    workbook_path = workbook_dir / "workbook.json"
    workbook_path.write_text(json.dumps(workbook, indent=2) + "\n", encoding="utf-8")
    generated(workbook_path)

    for sheet_name, sheet in workbook["sheets"].items():
        sheet_path = workbook_dir / f"{slug(sheet_name)}.json"
        sheet_path.write_text(
            json.dumps(
                {
                    "source": workbook["source"],
                    "workbook": workbook["workbook"],
                    **sheet,
                },
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        generated(sheet_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--file",
        action="append",
        dest="files",
        help="Workbook file name relative to --input-dir. Defaults to configured workbooks.",
    )
    parser.add_argument("--company-config", type=Path, default=DEFAULT_COMPANY_CONFIG)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = load_company_config(args.company_config)
    input_dir = args.input_dir or Path(config["inputDir"])
    files = tuple(args.files) if args.files else tuple(config.get("workbooks", []))
    if not files:
        raise ValueError(f"no workbooks configured in {config['_path']}")
    for file_name in files:
        path = input_dir / file_name
        if not path.is_file():
            raise FileNotFoundError(path)
        write_workbook_outputs(extract_workbook(path), args.output_dir)


if __name__ == "__main__":
    main()
